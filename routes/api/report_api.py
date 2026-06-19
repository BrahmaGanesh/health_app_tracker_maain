# ============================================================
# routes/api/report_api.py — Reports API (APK)
# PDF, Excel, Email reports
# ============================================================

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_current_user
from datetime import date, timedelta, datetime
import io

from extensions import db
from models import EmailReportConfig, EmailReportLog

report_api_bp = Blueprint("report_api", __name__)


def ok(data=None, msg="Success", code=200):
    r = {"success": True, "message": msg}
    if data is not None: r["data"] = data
    return jsonify(r), code


def err(msg="Error", code=400):
    return jsonify({"success": False, "message": msg}), code


# ── GET EMAIL REPORT CONFIG ───────────────────────────────────
@report_api_bp.route("/email-config", methods=["GET"])
@jwt_required()
def get_email_config():
    user   = get_current_user()
    config = user.email_report_config

    if not config:
        return ok({"configured": False, "config": None})

    return ok({
        "configured": True,
        "config": {
            "is_enabled":       config.is_enabled,
            "frequency":        config.frequency,
            "custom_days":      config.custom_days,
            "send_time":        config.send_time,
            "recipients":       config.recipient_list,
            "report_period_days": config.report_period_days,
            "include_weight":   config.include_weight,
            "include_bp":       config.include_bp,
            "include_water":    config.include_water,
            "include_sleep":    config.include_sleep,
            "include_steps":    config.include_steps,
            "include_exercise": config.include_exercise,
            "include_sugar":    config.include_sugar,
            "last_sent_at":     config.last_sent_at.isoformat() if config.last_sent_at else None,
            "total_sent":       config.total_sent,
        }
    })


# ── SAVE EMAIL REPORT CONFIG ──────────────────────────────────
@report_api_bp.route("/email-config", methods=["POST"])
@jwt_required()
def save_email_config():
    user   = get_current_user()
    data   = request.get_json() or {}
    config = user.email_report_config

    if not config:
        config = EmailReportConfig(user_id=user.id)
        db.session.add(config)

    config.is_enabled          = data.get("is_enabled", False)
    config.frequency           = data.get("frequency", "weekly")
    config.custom_days         = int(data.get("custom_days", 7))
    config.send_time           = data.get("send_time", "08:00")
    config.report_period_days  = int(data.get("report_period_days", 7))
    config.include_weight      = data.get("include_weight", True)
    config.include_bp          = data.get("include_bp", True)
    config.include_water       = data.get("include_water", True)
    config.include_sleep       = data.get("include_sleep", True)
    config.include_steps       = data.get("include_steps", True)
    config.include_exercise    = data.get("include_exercise", True)
    config.include_sugar       = data.get("include_sugar", True)
    config.include_suggestions = data.get("include_suggestions", True)

    # Recipients
    recipients = data.get("recipients", [])
    if isinstance(recipients, list):
        config.email_recipients = ", ".join(recipients)
    else:
        config.email_recipients = str(recipients)

    # Compute next send time
    from utils.email_sender import _compute_next_send
    if config.is_enabled:
        config.next_send_at = _compute_next_send(config)

    db.session.commit()
    return ok(message="Email report settings saved")


# ── SEND REPORT NOW ───────────────────────────────────────────
@report_api_bp.route("/send-now", methods=["POST"])
@jwt_required()
def send_now():
    user = get_current_user()
    data = request.get_json() or {}

    period_days = int(data.get("period_days", 7))
    recipients  = data.get("recipients", [user.email])

    try:
        from utils.email_sender import send_report_for_user
        success = send_report_for_user(user, period_days, recipients)
        if success:
            return ok(message=f"Report sent to {', '.join(recipients)}")
        return err("Failed to send report. Check email settings.", 500)
    except Exception as e:
        return err(f"Error: {str(e)}", 500)


# ── EMAIL REPORT HISTORY ──────────────────────────────────────
@report_api_bp.route("/history", methods=["GET"])
@jwt_required()
def report_history():
    user = get_current_user()
    logs = EmailReportLog.query.filter_by(user_id=user.id).order_by(
        EmailReportLog.sent_at.desc()
    ).limit(20).all()

    return ok({
        "logs": [
            {
                "id":           l.id,
                "sent_at":      l.sent_at.isoformat(),
                "recipients":   l.recipients,
                "period_start": str(l.period_start) if l.period_start else None,
                "period_end":   str(l.period_end) if l.period_end else None,
                "status":       l.status,
                "health_score": l.health_score,
            }
            for l in logs
        ]
    })


# ── GENERATE PDF REPORT ───────────────────────────────────────
@report_api_bp.route("/pdf", methods=["GET"])
@jwt_required()
def generate_pdf():
    user        = get_current_user()
    period_days = int(request.args.get("days", 7))
    today       = date.today()
    period_start= today - timedelta(days=period_days - 1)

    try:
        from utils.email_sender import _gather_report_data
        report_data = _gather_report_data(user, period_start, today)
        html        = _build_pdf_html(user, report_data, period_start, today)

        try:
            from weasyprint import HTML as WH
            pdf_bytes = WH(string=html).write_pdf()
            buf = io.BytesIO(pdf_bytes)
            buf.seek(0)
            filename = f"healthtrack_report_{today}.pdf"
            return send_file(buf, mimetype="application/pdf",
                             as_attachment=True, download_name=filename)
        except ImportError:
            # Fallback: return HTML
            return ok({
                "pdf_available": False,
                "message":       "WeasyPrint not installed. Email report available.",
                "html_preview":  html[:500],
            })

    except Exception as e:
        return err(f"PDF generation failed: {str(e)}", 500)


# ── GENERATE EXCEL REPORT ─────────────────────────────────────
@report_api_bp.route("/excel", methods=["GET"])
@jwt_required()
def generate_excel():
    user        = get_current_user()
    period_days = int(request.args.get("days", 7))
    today       = date.today()
    period_start= today - timedelta(days=period_days - 1)

    try:
        import openpyxl
        from utils.email_sender import _gather_report_data

        report_data = _gather_report_data(user, period_start, today)
        day_data    = report_data["day_data"]
        averages    = report_data["averages"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Health Report"

        # Style helpers
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill  = PatternFill("solid", fgColor="142D4C")
        header_font  = Font(color="FFFFFF", bold=True, size=11)
        accent_fill  = PatternFill("solid", fgColor="9FD3C7")
        accent_font  = Font(color="142D4C", bold=True)
        good_fill    = PatternFill("solid", fgColor="DCFCE7")
        bad_fill     = PatternFill("solid", fgColor="FEE2E2")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"🏥 HealthTrack Report — {user.name}"
        ws["A1"].font      = Font(size=16, bold=True, color="142D4C")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Period: {period_start.strftime('%d %b %Y')} to {today.strftime('%d %b %Y')} | Score: {report_data['health_score']}/100 ({report_data['grade']})"
        ws["A2"].font      = Font(size=11, color="4F3B78")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Headers row
        headers = ["Metric"] + [d["day_str"] for d in day_data] + ["Average"]
        row = 4
        for col, h in enumerate(headers, 1):
            cell         = ws.cell(row=row, column=col, value=h)
            cell.fill    = header_fill
            cell.font    = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border  = thin_border

        # Data rows
        metrics = [
            ("Weight (kg)",    "weight",  "kg"),
            ("BP Systolic",    "bp_sys",  "mmHg"),
            ("BP Diastolic",   "bp_dia",  "mmHg"),
            ("Water (L)",      "water",   "L"),
            ("Sleep (hrs)",    "sleep",   "hrs"),
            ("Steps",          "steps",   "steps"),
            ("Sugar (mg/dL)",  "sugar_f", "mg/dL"),
            ("Health Score",   "score",   "/100"),
        ]

        for r_offset, (label, key, unit) in enumerate(metrics):
            r = row + 1 + r_offset
            ws.cell(r, 1, value=f"{label} ({unit})").font = Font(bold=True)
            ws.cell(r, 1).fill   = accent_fill
            ws.cell(r, 1).font   = accent_font
            ws.cell(r, 1).border = thin_border

            for c_offset, d in enumerate(day_data):
                val  = d.get(key)
                cell = ws.cell(r, 2 + c_offset, value=val if val is not None else "—")
                cell.alignment = Alignment(horizontal="center")
                cell.border    = thin_border
                # Color coding for BP
                if key == "bp_sys" and val:
                    cell.fill = good_fill if val < 130 else (bad_fill if val >= 140 else PatternFill("solid", fgColor="FEF3C7"))

            # Average column
            avg_val = averages.get(key)
            avg_cell = ws.cell(r, 2 + len(day_data), value=avg_val if avg_val is not None else "—")
            avg_cell.font      = Font(bold=True)
            avg_cell.alignment = Alignment(horizontal="center")
            avg_cell.border    = thin_border

        # Achievements sheet
        ws2 = wb.create_sheet("Achievements")
        ws2["A1"] = "🏆 Achievements"
        ws2["A1"].font = Font(size=14, bold=True, color="142D4C")
        for i, ach in enumerate(report_data["achievements"], 2):
            ws2[f"A{i}"] = ach

        # Suggestions sheet
        ws3 = wb.create_sheet("Suggestions")
        ws3["A1"] = "💡 Personalised Suggestions"
        ws3["A1"].font = Font(size=14, bold=True, color="4F3B78")
        for i, sug in enumerate(report_data["suggestions"], 2):
            ws3[f"A{i}"] = sug

        # Auto-size columns
        for ws_item in [ws, ws2, ws3]:
            for col in ws_item.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                ws_item.column_dimensions[col_letter].width = min(max_len + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"healthtrack_report_{today}.xlsx"
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)

    except ImportError:
        return err("openpyxl not installed", 500)
    except Exception as e:
        return err(f"Excel generation failed: {str(e)}", 500)


def _build_pdf_html(user, data, period_start, period_end):
    from utils.email_sender import _generate_html_report
    return _generate_html_report(user, data, period_start, period_end)