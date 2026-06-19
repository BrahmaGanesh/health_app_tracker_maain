# routes/report_routes.py — PDF, Excel, CSV Export Website
import io
import openpyxl
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from datetime import date, timedelta

report_bp = Blueprint("reports", __name__)

@report_bp.route("/")
@login_required
def index():
    from models import EmailReportLog, EmailReportConfig
    logs   = EmailReportLog.query.filter_by(user_id=current_user.id).order_by(EmailReportLog.sent_at.desc()).limit(10).all()
    config = current_user.email_report_config
    return render_template("reports/index.html", logs=logs, config=config)

@report_bp.route("/pdf")
@login_required
def download_pdf():
    days        = int(request.args.get("days", 7))
    today       = date.today()
    period_start= today - timedelta(days=days - 1)
    try:
        from utils.email_sender import _gather_report_data, _generate_html_report
        data = _gather_report_data(current_user, period_start, today)
        html = _generate_html_report(current_user, data, period_start, today)
        try:
            from weasyprint import HTML as WH
            pdf   = WH(string=html).write_pdf()
            buf   = io.BytesIO(pdf); buf.seek(0)
            return send_file(buf, mimetype="application/pdf",
                             as_attachment=True, download_name=f"healthtrack_{today}.pdf")
        except ImportError:
            flash("PDF export requires WeasyPrint. Downloading HTML instead.", "warning")
            buf = io.BytesIO(html.encode()); buf.seek(0)
            return send_file(buf, mimetype="text/html",
                             as_attachment=True, download_name=f"healthtrack_{today}.html")
    except Exception as e:
        flash(f"Export failed: {e}", "danger")
        return redirect(url_for("reports.index"))

@report_bp.route("/excel")
@login_required
def download_excel():
    days        = int(request.args.get("days", 7))
    today       = date.today()
    period_start= today - timedelta(days=days - 1)
    try:
        from routes.api.report_api import _build_pdf_html
        from utils.email_sender import _gather_report_data
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        data     = _gather_report_data(current_user, period_start, today)
        day_data = data["day_data"]
        averages = data["averages"]

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Health Report"
        hf = PatternFill("solid", fgColor="142D4C")
        af = PatternFill("solid", fgColor="9FD3C7")
        tb = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

        ws.merge_cells("A1:H1")
        ws["A1"] = f"HealthTrack Report — {current_user.name} | {period_start.strftime('%d %b')} to {today.strftime('%d %b %Y')} | Score: {data['health_score']}/100"
        ws["A1"].font = Font(size=14,bold=True,color="142D4C")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Metric"] + [d["day_str"] for d in day_data] + ["Average"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(4, col, value=h); c.fill = hf; c.font = Font(color="FFFFFF",bold=True); c.border = tb; c.alignment = Alignment(horizontal="center")

        rows = [("Weight (kg)","weight"),("BP Systolic","bp_sys"),("BP Diastolic","bp_dia"),
                ("Water (L)","water"),("Sleep (hrs)","sleep"),("Steps","steps"),("Sugar","sugar_f"),("Score","score")]
        for ri, (label, key) in enumerate(rows):
            r = 5 + ri
            lc = ws.cell(r,1,value=label); lc.fill=af; lc.font=Font(bold=True,color="142D4C"); lc.border=tb
            for ci, d in enumerate(day_data):
                c = ws.cell(r, 2+ci, value=d.get(key)); c.alignment=Alignment(horizontal="center"); c.border=tb
            avg_c = ws.cell(r, 2+len(day_data), value=averages.get(key)); avg_c.font=Font(bold=True); avg_c.alignment=Alignment(horizontal="center"); avg_c.border=tb

        ws2 = wb.create_sheet("Achievements")
        ws2["A1"] = "🏆 Achievements"; ws2["A1"].font = Font(size=12,bold=True)
        for i, a in enumerate(data["achievements"],2): ws2[f"A{i}"] = a

        ws3 = wb.create_sheet("Suggestions")
        ws3["A1"] = "💡 Suggestions"; ws3["A1"].font = Font(size=12,bold=True)
        for i, s in enumerate(data["suggestions"],2): ws3[f"A{i}"] = s

        for ws_item in [ws, ws2, ws3]:
            for col in ws_item.columns:
                ml = max((len(str(c.value or "")) for c in col), default=10)
                ws_item.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 30)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"healthtrack_{today}.xlsx")
    except ImportError:
        flash("openpyxl required for Excel export. Install with: pip install openpyxl", "warning")
        return redirect(url_for("reports.index"))
    except Exception as e:
        flash(f"Export failed: {e}", "danger")
        return redirect(url_for("reports.index"))

@report_bp.route("/csv")
@login_required
def download_csv():
    import csv, io as sio
    days        = int(request.args.get("days", 7))
    today       = date.today()
    period_start= today - timedelta(days=days - 1)
    from utils.email_sender import _gather_report_data
    data     = _gather_report_data(current_user, period_start, today)
    day_data = data["day_data"]

    output = sio.StringIO()
    w = csv.writer(output)
    w.writerow(["Date","Weight(kg)","BP Sys","BP Dia","Water(L)","Sleep(hrs)","Steps","Sugar(mg/dL)","Score"])
    for d in day_data:
        w.writerow([d["date"],d["weight"] or "",d["bp_sys"] or "",d["bp_dia"] or "",
                    d["water"] or "",d["sleep"] or "",d["steps"] or "",d["sugar_f"] or "",d["score"] or ""])
    w.writerow([])
    w.writerow(["Averages","","","","","","","",""])
    avgs = data["averages"]
    w.writerow(["Average",avgs.get("weight",""),avgs.get("bp_sys",""),avgs.get("bp_dia",""),
                avgs.get("water",""),avgs.get("sleep",""),avgs.get("steps",""),"",avgs.get("score","")])

    buf = io.BytesIO(output.getvalue().encode()); buf.seek(0)
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name=f"healthtrack_{today}.csv")

